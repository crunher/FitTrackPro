#!/usr/bin/env python3
"""
Exercise Import Script for FitTrack Pro
Parses exercises.xlsx and generates exercises.json for Android app seeding
"""

import json
import re
import openpyxl
from pathlib import Path

# Configuration - Supabase Storage URL
SUPABASE_URL = "https://yflyygasrbzzsjolosjp.supabase.co/storage/v1/object/public/exercises"

# Category mapping: Polish -> ExerciseCategory enum
CATEGORY_MAP = {
    "barki": "SHOULDERS",
    "biceps": "ARMS",
    "brzuch": "ABS",
    "crossfit": "COMPOUND",
    "cwiczenia-dla-kobiet": "BODYWEIGHT",
    "cwiczenia-domowe": "BODYWEIGHT",
    "czworoglowe-uda": "LEGS",
    "dwuglowe-uda": "LEGS",
    "dwuglowe-uda-posladki": "LEGS",
    "kettlebell": "KETTLEBELL",
    "klatka-piersiowa": "CHEST",
    "lydki": "LEGS",
    "mobilizacje": "STRETCHING",
    "plecy": "BACK",
    "rolowanie": "STRETCHING",
    "street-workout": "BODYWEIGHT",
    "triceps": "ARMS",
}

# Muscle mapping: Polish -> MuscleGroup enum
MUSCLE_MAP = {
    # Shoulders
    "naramienne": "SHOULDERS",
    "naramienny": "SHOULDERS",
    "stożka rotatorów": "SHOULDERS",
    "stożek rotatorów": "SHOULDERS",
    "rotatorów": "SHOULDERS",
    "barki": "SHOULDERS",
    
    # Back
    "najszerszy grzbietu": "BACK",
    "najszersze grzbietu": "BACK",
    "czworoboczny": "TRAPS",
    "czworoboczne": "TRAPS",
    "prostownik grzbietu": "LOWER_BACK",
    "prostowniki grzbietu": "LOWER_BACK",
    "równoległoboczny": "BACK",
    "równoległoboczne": "BACK",
    "grzbiet": "BACK",
    "plecy": "BACK",
    
    # Chest
    "piersiowy": "CHEST",
    "piersiowe": "CHEST",
    "klatka piersiowa": "CHEST",
    
    # Arms
    "biceps": "BICEPS",
    "dwugłowy ramienia": "BICEPS",
    "triceps": "TRICEPS",
    "trójgłowy ramienia": "TRICEPS",
    "przedramion": "FOREARMS",
    "przedramiona": "FOREARMS",
    
    # Core
    "prosty brzucha": "ABS",
    "skośny brzucha": "OBLIQUES",
    "skośne brzucha": "OBLIQUES",
    "brzuch": "ABS",
    
    # Legs
    "czworogłowy uda": "QUADS",
    "czworogłowe uda": "QUADS",
    "czworogłowy": "QUADS",
    "dwugłowy uda": "HAMSTRINGS",
    "kulszowo-goleniowy": "HAMSTRINGS",
    "kulszowo-goleniowe": "HAMSTRINGS",
    "pośladkowy": "GLUTES",
    "pośladkowe": "GLUTES",
    "pośladki": "GLUTES",
    "łydki": "CALVES",
    "łydka": "CALVES",
    "przywodziciele": "ADDUCTORS",
    "odwodziciele": "ABDUCTORS",
    
    # Other
    "serce": "CARDIO",
}


def extract_muscles(description: str) -> tuple[str, list[str]]:
    """Extract muscles from description text."""
    if not description:
        return "OTHER", []
    
    # Find the "Mięśnie zaangażowane:" section
    muscle_section = ""
    match = re.search(r'Mięśnie zaangażowane:(.*?)(?:Pozycja wyjściowa|$)', 
                      description, re.DOTALL | re.IGNORECASE)
    if match:
        muscle_section = match.group(1).lower()
    else:
        # Try first few lines
        muscle_section = description[:500].lower()
    
    main_muscle = "OTHER"
    secondary_muscles = []
    
    # Find all matching muscles
    found_muscles = []
    for polish_name, enum_value in MUSCLE_MAP.items():
        if polish_name.lower() in muscle_section:
            if enum_value not in [m[1] for m in found_muscles]:
                found_muscles.append((polish_name, enum_value))
    
    if found_muscles:
        main_muscle = found_muscles[0][1]
        secondary_muscles = [m[1] for m in found_muscles[1:4]]  # Max 3 secondary
    
    return main_muscle, secondary_muscles


def parse_description(description: str) -> dict:
    """Parse description into structured sections."""
    if not description:
        return {"muscles": [], "position": [], "movement": [], "tips": []}
    
    result = {
        "muscles": [],
        "position": [],
        "movement": [],
        "tips": []
    }
    
    # Extract muscles section
    muscles_match = re.search(r'Mięśnie zaangażowane:(.*?)(?:Pozycja|$)', 
                              description, re.DOTALL | re.IGNORECASE)
    if muscles_match:
        muscles_text = muscles_match.group(1).strip()
        result["muscles"] = [m.strip().strip('–').strip('•').strip() 
                            for m in re.split(r'[,\n]', muscles_text) 
                            if m.strip() and len(m.strip()) > 2]
    
    # Extract position section
    pos_match = re.search(r'Pozycja wyjściowa(.*?)(?:Ruch|$)', 
                          description, re.DOTALL | re.IGNORECASE)
    if pos_match:
        result["position"] = extract_numbered_list(pos_match.group(1))
    
    # Extract movement section
    mov_match = re.search(r'Ruch(.*?)(?:Wskazówki|$)', 
                          description, re.DOTALL | re.IGNORECASE)
    if mov_match:
        result["movement"] = extract_numbered_list(mov_match.group(1))
    
    # Extract tips section
    tips_match = re.search(r'Wskazówki(.*?)$', 
                           description, re.DOTALL | re.IGNORECASE)
    if tips_match:
        result["tips"] = extract_numbered_list(tips_match.group(1))
    
    return result


def extract_numbered_list(text: str) -> list[str]:
    """Extract numbered items from text."""
    items = []
    for match in re.finditer(r'\d+\)\s*(.+?)(?=\d+\)|$)', text, re.DOTALL):
        item = match.group(1).strip()
        if item and len(item) > 5:
            items.append(item)
    return items


def parse_images(images_str: str) -> list[str]:
    """Parse semicolon-separated image paths to URLs."""
    if not images_str:
        return []
    
    images = []
    for path in images_str.split(';'):
        path = path.strip().replace('\\', '/')
        if path:
            # Convert local path to Supabase URL
            url = f"{SUPABASE_URL}/{path}"
            images.append(url)
    return images


def parse_video(video_path: str) -> str | None:
    """Convert video path to Supabase URL."""
    if not video_path:
        return None
    
    path = video_path.strip().replace('\\', '/')
    return f"{SUPABASE_URL}/{path}"


def determine_tracking_type(category: str, name: str) -> str:
    """Determine tracking type based on category and name."""
    name_lower = name.lower() if name else ""
    
    # Cardio exercises
    if category in ["crossfit"] or "bieg" in name_lower or "marsz" in name_lower:
        return "DISTANCE_X_KCAL_X_TIME"
    
    # Time-based
    if "deska" in name_lower or "plank" in name_lower or "hollow" in name_lower:
        return "TIME"
    
    # Stretching/mobility
    if category in ["mobilizacje", "rolowanie"]:
        return "TIME"
    
    # Default to weight x reps
    return "WEIGHT_X_REPS"


def is_unilateral(name: str) -> bool:
    """Check if exercise is unilateral (one-sided)."""
    unilateral_keywords = [
        "jednorącz", "jednonóż", "jednoręcz", "jednonożn",
        "jednoracz", "jednonoz", "jednostronn",
        "one arm", "single arm", "single leg"
    ]
    name_lower = name.lower() if name else ""
    return any(kw in name_lower for kw in unilateral_keywords)


def main():
    print("Loading exercises.xlsx...")
    wb = openpyxl.load_workbook('exercises.xlsx')
    ws = wb.active
    
    exercises = []
    
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 1).value
        category_pl = ws.cell(row_idx, 2).value
        description = ws.cell(row_idx, 3).value
        video_path = ws.cell(row_idx, 4).value
        images_str = ws.cell(row_idx, 5).value
        
        if not name:
            continue
        
        # Map category
        category = CATEGORY_MAP.get(category_pl, "OTHER")
        
        # Extract muscles
        main_muscle, secondary_muscles = extract_muscles(description)
        
        # Parse structured description
        structured_desc = parse_description(description)
        
        # Build exercise object
        exercise = {
            "id": row_idx - 1,  # 1-based ID
            "name": name.strip(),
            "namePl": name.strip(),  # Already in Polish
            "description": description.strip() if description else None,
            "descriptionStructured": structured_desc,
            "category": category,
            "trackingType": determine_tracking_type(category_pl, name),
            "mainMuscle": main_muscle,
            "secondaryMuscles": secondary_muscles,
            "videoUrl": parse_video(video_path),
            "imageUrls": parse_images(images_str),
            "unilateral": is_unilateral(name),
            "volumeMultiplier": 2.0 if is_unilateral(name) else 1.0,
            "isCustom": False
        }
        
        exercises.append(exercise)
        
        if row_idx % 50 == 0:
            print(f"Processed {row_idx - 1} exercises...")
    
    # Save to JSON
    output_path = Path("app/src/main/assets/exercises.json")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(exercises, f, ensure_ascii=False, indent=2)
    
    print(f"\n[SUCCESS] Exported {len(exercises)} exercises to {output_path}")
    
    # Print statistics
    categories = {}
    muscles = {}
    for ex in exercises:
        cat = ex["category"]
        categories[cat] = categories.get(cat, 0) + 1
        muscle = ex["mainMuscle"]
        muscles[muscle] = muscles.get(muscle, 0) + 1
    
    print("\n[STATS] Categories distribution:")
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")
    
    print("\n[STATS] Main muscles distribution:")
    for muscle, count in sorted(muscles.items(), key=lambda x: -x[1]):
        print(f"  {muscle}: {count}")


if __name__ == "__main__":
    main()
